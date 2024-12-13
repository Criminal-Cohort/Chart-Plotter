from datetime import datetime, timedelta
import os
import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
import pandas as pd
from dateutil.relativedelta import relativedelta
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import yfinance as yf
from openpyxl.styles import NamedStyle, PatternFill, Border, Side
from pathlib import Path

# Create the main window
app = tk.Tk()
app.title("Folio Manager")

# Set the initial window size and allow it to be scalable
app.geometry("550x500")
app.minsize(400, 300)

# Create a notebook (tab holder)
tab_notebook = ttk.Notebook(app)
tab_notebook.pack(fill='both', expand=True)

# Define font styles
default_font = ("Helvetica", 14)
header_font = ("Arial", 16)

# Tab 1: Find Live
find_live_tab = ttk.Frame(tab_notebook)
tab_notebook.add(find_live_tab, text='Find Live')

# Create frames for better organization
top_frame_find_live = ttk.Frame(find_live_tab, padding=10)
top_frame_find_live.grid(row=0, column=0, columnspan=2, sticky="ew")
low_frame_find_live = ttk.Frame(find_live_tab, padding=10)
low_frame_find_live.grid(row=1, column=0, columnspan=2, sticky="ew")

# Configure resizing
find_live_tab.columnconfigure(0, weight=1)
find_live_tab.rowconfigure(2, weight=1)

# Widgets in the Find Live tab
label_ticker = ttk.Label(top_frame_find_live, text="Ticker:", font=default_font)
label_ticker.grid(row=0, column=0, padx=(0, 10), sticky="e")

entry_ticker = ttk.Entry(top_frame_find_live, font=default_font, width=20)
entry_ticker.grid(row=0, column=1, sticky="w")

checkbox_var = tk.IntVar()
check_find_closed = ttk.Checkbutton(low_frame_find_live, text="Find Closed", variable=checkbox_var)
check_find_closed.grid(row=0, column=0, sticky="w")

label_days = ttk.Label(low_frame_find_live, text="Days:", font=default_font)
label_days.grid(row=0, column=1, padx=(20, 5), sticky="e")

entry_days = ttk.Entry(low_frame_find_live, font=default_font, width=10)
entry_days.grid(row=0, column=2, sticky="w")

find_button = ttk.Button(low_frame_find_live, text="Find", command=lambda: handle_find_live())
find_button.grid(row=0, column=3, padx=(20, 0))

# Add a text area with a scrollbar
text_area_find_live = tk.Text(find_live_tab, font=("Courier", 10), wrap="word")
text_area_find_live.grid(row=2, column=0, columnspan=2, sticky="nsew", padx=10, pady=10)

scrollbar_find_live = ttk.Scrollbar(find_live_tab, orient="vertical", command=text_area_find_live.yview)
text_area_find_live.configure(yscrollcommand=scrollbar_find_live.set)
scrollbar_find_live.grid(row=2, column=2, sticky="ns")

# Tab 2: Apply RS
apply_rs_tab = ttk.Frame(tab_notebook)
tab_notebook.add(apply_rs_tab, text='Apply RS')

# Create top frame in Apply RS tab
top_frame_apply_rs = tk.Frame(apply_rs_tab)
top_frame_apply_rs.grid(row=0, column=0, columnspan=3, padx=10, pady=10, sticky='ew')

# Configure grid columns for the top frame
top_frame_apply_rs.grid_columnconfigure(0, weight=1)
top_frame_apply_rs.grid_columnconfigure(1, weight=1)
top_frame_apply_rs.grid_columnconfigure(2, weight=1)

# Create ticker frame
ticker_frame = tk.Frame(top_frame_apply_rs)
ticker_frame.grid(row=0, column=0, columnspan=3, pady=10)

label_ticker_apply = ttk.Label(ticker_frame, text="Ticker:", font=default_font)
label_ticker_apply.grid(row=0, column=0, padx=(0, 10), pady=10, sticky='e')

entry_ticker_apply = ttk.Entry(ticker_frame, font=default_font, width=15)
entry_ticker_apply.grid(row=0, column=1, padx=(0, 10), pady=10, sticky='w')

# Create split frame
split_frame = tk.Frame(top_frame_apply_rs)
split_frame.grid(row=1, column=0, columnspan=2, pady=10)

label_split_coefficient = ttk.Label(split_frame, text="Split Coeff:", font=default_font)
label_split_coefficient.grid(row=0, column=0, padx=(0, 10), pady=10, sticky='e')

entry_split_coefficient = ttk.Entry(split_frame, font=default_font, width=5)
entry_split_coefficient.grid(row=0, column=1, padx=(0, 10), pady=10, sticky='w')

apply_button = tk.Button(split_frame, text="Apply", font=default_font, width=10, command=lambda: handle_apply_rs())
apply_button.grid(row=0, column=2, padx=(10, 0), pady=10, sticky='w')

text_area_apply_rs = tk.Text(apply_rs_tab, font=default_font)
text_area_apply_rs.grid(row=3, column=0, columnspan=3, padx=10, pady=10, sticky='nsew')

apply_rs_tab.columnconfigure(0, weight=1)
apply_rs_tab.rowconfigure(2, weight=1)

# Tab 3: Find Comment
find_comment_tab = ttk.Frame(tab_notebook)
tab_notebook.add(find_comment_tab, text='Find Comment')

top_frame_find_comment = ttk.Frame(find_comment_tab)
top_frame_find_comment.grid(row=0, column=0, columnspan=2, padx=10, pady=10, sticky='ew')

top_frame_find_comment.grid_columnconfigure(0, weight=1)
top_frame_find_comment.grid_columnconfigure(1, weight=1)
find_comment_tab.grid_columnconfigure(0, weight=1)
find_comment_tab.grid_columnconfigure(1, weight=1)

label_comment = ttk.Label(top_frame_find_comment, text="Comment:", font=default_font)
label_comment.grid(row=0, column=0, padx=(0, 10), pady=10, sticky='e')

entry_comment = ttk.Entry(top_frame_find_comment, font=default_font, width=15)
entry_comment.grid(row=0, column=1, padx=(10, 0), pady=10, sticky='w')

label_ticker_comment = ttk.Label(top_frame_find_comment, text="Ticker:", font=default_font)
label_ticker_comment.grid(row=1, column=0, padx=(0, 10), pady=10, sticky='e')

entry_ticker_comment = ttk.Entry(top_frame_find_comment, font=default_font, width=15)
entry_ticker_comment.grid(row=1, column=1, padx=(10, 0), pady=10, sticky='w')

find_button_comment = tk.Button(find_comment_tab, text="Find", font=default_font, width=14,
                                command=lambda: handle_find_comment())
find_button_comment.grid(row=2, column=0, padx=(10, 0), pady=10, sticky='e')

text_area_find_comment = tk.Text(find_comment_tab, font=default_font)
text_area_find_comment.grid(row=3, column=0, columnspan=2, padx=10, pady=10, sticky='nsew')

# Tab 4: Settings
settings_tab = ttk.Frame(tab_notebook)
tab_notebook.add(settings_tab, text='Settings')

label_file_path = tk.Label(settings_tab, text="Folios Folder Path:", font=header_font)
label_file_path.pack(pady=20, padx=10)

file_path_var = tk.StringVar(value="C:\\Program Files (x86)\\RadiantInvestments\\folioTrackerSetup\\folios")
entry_file_path = tk.Entry(settings_tab, width=52, font=default_font, textvariable=file_path_var, highlightthickness=1,
                           highlightbackground="black", highlightcolor="black")
entry_file_path.pack(pady=20, padx=20)

refresh_button = tk.Button(settings_tab, text="Cleanup Folios", font=default_font, command=lambda: handle_cleaup())
refresh_button.pack(pady=20, padx=20, anchor='e')

seller_button = tk.Button(settings_tab, text="Process ForcedSell", font=default_font, command=lambda: handle_sellem())
seller_button.pack(pady=20, padx=20, anchor='e')


def find_files(folder_path):
    try:
        files = [file for file in os.listdir(folder_path) if file.endswith(".xlsx")]
        return files
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")
        return []


def handle_find_live():
    folder_path = file_path_var.get()
    if not os.path.isdir(folder_path):
        messagebox.showerror("Error", "Invalid Folios Folder path")
        return

    find_files(folder_path)
    if checkbox_var.get():
        process_files(folder_path, extra=True)
    else:
        process_files(folder_path)


def process_files(folder_path, extra=False):
    global files_inserted

    files_inserted = []
    files = find_files(folder_path)
    text_area_find_live.delete(1.0, tk.END)

    for file in files:
        if file.startswith("~"):
            continue

        file_path = os.path.join(folder_path, file)

        try:
            with open(file_path, 'r+b'):
                pass
        except IOError:
            messagebox.showwarning("Warning", f"Couldn't process {file} as it was open.")
            continue

        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active
            results = ""
            final_day = datetime.today() - timedelta(days=int(entry_days.get()))

            for idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=1):
                if extra:
                    if len(row) > 12 and (row[0] or row[11] or row[12]):
                        if row[1] == entry_ticker.get():
                            print(row[2], final_day)
                            if datetime.strptime(row[2], "%d-%b-%Y").date() >= final_day.date():
                                results += f"{idx + 2}, "

                else:
                    if len(row) > 1 and row[0] and row[1] == entry_ticker.get().upper():
                        results += f"{idx + 2}, "

            if results:
                text_area_find_live.insert(tk.END, f"{file[:-11]} in row {results[:-2]}\n")
                files_inserted.append(file)

        except Exception as e:
            text_area_find_live.insert(tk.END, f"Error reading file {file}: {e}\n")

    if not text_area_find_live.get("1.0", tk.END).strip():
        text_area_find_live.insert(tk.END, f"No valid rows for {entry_ticker.get()}.\n")


def handle_apply_rs():
    split_coefficient = entry_split_coefficient.get()
    if split_coefficient:
        apply_split_coefficient(file_path_var.get(), split_coefficient)
    else:
        messagebox.showerror("Error", "Please enter a split coefficient value.")


def apply_split_coefficient(folder_path, split_coefficient):
    global files_edited

    files_edited = []
    files = find_files(folder_path)
    text_area_apply_rs.delete(1.0, tk.END)

    for file in files:
        if file.startswith("~"):
            continue

        file_path = os.path.join(folder_path, file)

        try:
            with open(file_path, 'r+b'):
                pass
        except IOError:
            messagebox.showwarning("Warning", f"Couldn't process {file} as it was open.")
            continue

        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active
            edited = 0

            for idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=1):
                if len(row) > 1 and row[0] and row[1] == entry_ticker_apply.get().upper():
                    sheet[f"D{idx + 2}"] = f"={(sheet[f'D{idx + 2}'].value)}*{split_coefficient}"
                    sheet[f"E{idx + 2}"] = f"=ROUNDDOWN({(sheet[f'E{idx + 2}'].value)}/{split_coefficient}, 0)"
                    sheet[f"F{idx + 2}"] = f"={(sheet[f'F{idx + 2}'].value)}*{split_coefficient}"
                    sheet[f"H{idx + 2}"] = f"={(sheet[f'H{idx + 2}'].value)}*{split_coefficient}"
                    sheet[
                        f"N{idx + 2}"] = f"{split_coefficient}:1 Reverse Split Dated: {datetime.today().strftime('%m-%d-%Y')}, " + f"{sheet[f'N{idx + 2}'].value}"
                    edited += 1

            if edited:
                workbook.save(file_path)
                text_area_apply_rs.insert(tk.END, f"Edits made in {file[:-11]}\n")
                files_edited.append(file)

        except Exception as e:
            text_area_apply_rs.insert(tk.END, f"Error reading file {file}: {e}\n")

    if not text_area_apply_rs.get("1.0", tk.END).strip():
        text_area_apply_rs.insert(tk.END, f"No valid rows for {entry_ticker_apply.get()}.\n")


def handle_find_comment():
    if entry_ticker_comment.get():
        find_comments(file_path_var.get(), ticker=entry_ticker_comment.get())
    else:
        find_comments(file_path_var.get())


def find_comments(folder_path, ticker=None):
    files = find_files(folder_path)
    text_area_find_comment.delete(1.0, tk.END)

    for file in files:
        file_path = os.path.join(folder_path, file)

        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active

            for idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=1):
                if entry_comment.get() in row[13] and (ticker is None or row[1] == ticker):
                    pass

        except Exception as e:
            text_area_find_comment.insert(tk.END, f"Error reading file {file}: {e}\n")


def clean_up(folder_path):
    files = find_files(folder_path)
    eight_months_ago = datetime.today() - relativedelta(months=6)

    for file in files:
        if file.startswith("~"):
            continue

        file_path = os.path.join(folder_path, file)

        try:
            with open(file_path, 'r+b'):
                pass
        except IOError:
            messagebox.showwarning("Warning", f"Couldn't process {file} as it was open.")
            continue

        try:
            # Read the Excel file with pandas
            df = pd.read_excel(file_path, sheet_name=0, header=None)

            workbook = load_workbook(file_path)
            sheet = workbook.active
            dated = []
            r5 = []
            r9 = []
            r10 = []
            r11 = []
            r15 = []
            r16 = []

            for idx, row in df.iterrows():
                r5.append(row[4])
                r9.append(row[8])
                r10.append(row[9])
                r11.append(row[10])
                r15.append(row[14])
                r16.append(row[15])

                if len(row) > 1 and pd.isna(row[0]) and not pd.isna(row[1]) and pd.isna(row[11]) and pd.isna(row[12]):
                    print(int(str(idx)) + 2, row[1])
                    date_str = row[2]
                    date_use = str(date_str).strip(" ")
                    print(date_use)
                    print("222222")
                    try:
                        date = datetime.strptime(date_use, "%d-%b-%Y")
                    except:
                        date = datetime.strptime(date_use, "%Y-%m-%d %H:%M:%S")
                    print(date)
                    if date < eight_months_ago:
                        print("hehehe")
                        print(int(str(idx)) + 2, file_path)
                        dated.append(int(str(idx)) + 2)

            # Delete rows from the actual Excel sheet using openpyxl
            dated = dated[::-1]
            for i in dated:
                print("deleting")
                sheet.delete_rows(i)

            # Save the modified Excel file
            workbook.save(file_path)

        except Exception as e:
            messagebox.showerror("Error", f"Error reading {file}:\n{e}\n")

    messagebox.showinfo("Finished", "Cleaned up Folios!")


def handle_cleaup():
    clean_up(entry_file_path.get())


def color_code(use_sheet, color, row):
    if color == 'b':
        for i in range(ord("B"), ord("P")+1):
            use_sheet[f"{chr(i)}{row}"].fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')

    elif color == "r":
        for i in range(ord("B"), ord("P") + 1):
            use_sheet[f"{chr(i)}{row}"].fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    elif color == 'y':
        for i in range(ord("B"), ord("P") + 1):
            use_sheet[f"{chr(i)}{row}"].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')


def sellem(folder_path, worker_sheet):
    files = find_files(folder_path)
    print(files)

    try:
        with open(str(os.getcwd()) + "\\Unprocessed_ForcedSells.xlsx", 'r+b'):
            pass
    except IOError:
        messagebox.showwarning("Warning", f"Couldn't process Unprocessed_ForcedSells as it was open.")

    for file in files:
        file_path = os.path.join(folder_path, file)
        print(file, file_path)

        """try:
            with open(file_path, 'r+b'):
                pass
        except IOError:
            messagebox.showwarning("Warning", f"Couldn't process {file} as it was open.")
            continue"""

        try:
            workbook = load_workbook(file_path)
            worksheet = workbook.active

            for idx, row in enumerate(worksheet.iter_rows(values_only=True, min_row=3), start=3):
                try:
                    if worksheet[f"A{idx}"].value:
                        today = datetime.now()
                        ticker = worksheet[f"B{idx}"].value
                        ticker_data = yf.Ticker(ticker)
                        close_val = ticker_data.history(period="1d")['Close'].iloc[0]
                        close_val = round(close_val, 2)
                        y = False

                        if worksheet[f"G{idx}"].value and worksheet[f"H{idx}"].value:
                            if str(worksheet[f"I{idx}"].value)[0] == "=":
                                if str(worksheet[f"E{idx}"].value)[0] == "=":
                                    e_vals = str(worksheet[f"E{idx}"].value).split(",")
                                    i_val = round(eval(e_vals[0][11:]), 0)
                                else:
                                    i_val = worksheet[f"E{idx}"].value
                            else:
                                i_val = worksheet[f"I{idx}"].value

                            if str(worksheet[f"H{idx}"].value)[0] == "=":
                                h_val = eval(str(worksheet[f"H{idx}"].value)[1:])
                            else:
                                h_val = worksheet[f"H{idx}"].value

                            i_val = float(i_val)
                            h_val = float(h_val)

                            if h_val == close_val:
                                color_code(worksheet, 'b', idx)

                            elif h_val > close_val:
                                color_code(worksheet, 'r', idx)

                            elif h_val < close_val:
                                color_code(worksheet, 'y', idx)
                                y = True

                            worksheet[f"M{idx}"].value = close_val
                            worksheet[f"L{idx}"].value = datetime.today().strftime("%d-%b-%y")
                            worksheet[f"A{idx}"].value = ""

                        if not y:
                            worksheet[f"N{idx}"].value = f"ForcedSell {today.strftime("%b")}{today.strftime("%y")}SentimentBelow60, " + worksheet[f"N{idx}"].value

                        else:
                            worksheet[
                                f"N{idx}"].value = f"ForcedSell U4 Stopped ${int(((close_val-h_val)*i_val)//1)} ie {int((((close_val-h_val)*i_val)*100/i_val*h_val)//1)} %, " + worksheet[f"N{idx}"].value

                except IndexError:
                    worker_sheet.insert_rows(idx=1)

                    worker_sheet.cell(row=1, column=1).value = ticker
                    worker_sheet.cell(row=1, column=2).value = file.strip(".xlsx")

                    continue

            workbook.save(file_path)


        except Exception as e:
            messagebox.showerror("Error", f"Error reading {file}:\n{e}\n")

    messagebox.showinfo("Finished", "ForcedSell Applied!")


def handle_sellem():
    worker = load_workbook(str(os.getcwd()) + "\\Unprocessed_ForcedSells.xlsx")
    worker_sheet = worker.active

    for row in worker_sheet.iter_rows():
        for cell in row:
            cell.value = None

    sellem(entry_file_path.get(), worker_sheet)

    worker.save(str(os.path.dirname(os.path.abspath(__file__))) + "\\Unprocessed_ForcedSells.xlsx")

    os.startfile(str(os.path.dirname(os.path.abspath(__file__))) + "\\Unprocessed_ForcedSells.xlsx")


def on_text_click(event):
    try:
        index = text_area_find_live.index('@%s,%s' % (event.x, event.y))
        line_number = int(index.split('.')[0])
        file_name = files_inserted[line_number - 1]

        if file_name:
            os.startfile(os.path.join(entry_file_path.get(), file_name))
    except Exception as e:
        print(f"Error handling click: {e}")


def on_text_click_edits(event):
    try:
        index = text_area_apply_rs.index('@%s,%s' % (event.x, event.y))
        line_number = int(index.split('.')[0])
        file_name = files_edited[line_number - 1]

        if file_name:
            os.startfile(os.path.join(entry_file_path.get(), file_name))
    except Exception as e:
        print(f"Error handling click: {e}")


# Bind the text area to the click event
text_area_find_live.bind("<Double-1>", on_text_click)
text_area_apply_rs.bind("<Double-1>", on_text_click_edits)

# Start the main event loop
app.mainloop()
